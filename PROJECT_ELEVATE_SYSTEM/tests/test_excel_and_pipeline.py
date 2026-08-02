# -*- coding: utf-8 -*-
"""Tests for the Excel generator and the end-to-end pipeline.
اختبارات مولّد الإكسل والمنسّق الكامل."""
import pytest
from openpyxl import load_workbook

from conftest import RATES, MODULE_DIR
from export_excel_template import ElevateWorkbookBuilder
from gainsharing_calculator import GainsharingCalculator, ProjectFinancials, TeamMember


EXPECTED_SHEETS = ["Gainsharing", "Pool & Cash Gate", "BOQ Audit", "Site KPIs", "Governance"]


def _live_result():
    calc = GainsharingCalculator(rates_path=RATES)
    fin = ProjectFinancials(
        project_name="T", baseline_cost_egp=12_000_000.0, actual_cost_egp=10_400_000.0,
        cash_collected_pct=0.82, quality_factor=0.95, escalation_commodity="steel_rebar",
        bad_debt_egp=150_000.0, subcontractor_value_egp=3_000_000.0, lost_time_injuries=0,
    )
    members = [
        TeamMember("Ahmed", "Site Manager", ld_badge="Level 3", ppc=0.92,
                   equipment_oee=0.97, vo_settlement_days=5,
                   value_engineering_savings_egp=200_000),
        TeamMember("Sara", "Planner", ld_badge="Level 2", ppc=0.90, time_weight=0.5),
    ]
    return calc.run(fin, members)


def test_template_workbook_builds(tmp_path):
    out = tmp_path / "template.xlsx"
    ElevateWorkbookBuilder(rates_path=RATES).build(out)
    assert out.exists()
    wb = load_workbook(out)
    assert wb.sheetnames == EXPECTED_SHEETS


def test_template_base_perf_cells_are_zero(tmp_path):
    out = tmp_path / "template.xlsx"
    ElevateWorkbookBuilder(rates_path=RATES).build(out)
    gs = load_workbook(out)["Gainsharing"]
    # First data row is 5; Base col=9, Perf col=10.
    assert gs.cell(5, 9).value == 0
    assert gs.cell(5, 10).value == 0


def test_live_workbook_seeds_engine_values(tmp_path):
    result = _live_result()
    out = tmp_path / "live.xlsx"
    ElevateWorkbookBuilder(rates_path=RATES, gainsharing_result=result).build(out)
    gs = load_workbook(out)["Gainsharing"]
    dist = result.distribution_df.set_index("name")
    # Ahmed's seeded Base share must match the engine to the cent.
    assert gs.cell(5, 1).value == "Ahmed"
    assert gs.cell(5, 9).value == pytest.approx(dist.loc["Ahmed", "base_share_egp"])


def test_live_workbook_keeps_live_formulas(tmp_path):
    result = _live_result()
    out = tmp_path / "live.xlsx"
    ElevateWorkbookBuilder(rates_path=RATES, gainsharing_result=result).build(out)
    gs = load_workbook(out)["Gainsharing"]
    # Gross (col 11) must remain a SUM formula, not a hard number.
    assert str(gs.cell(5, 11).value).startswith("=SUM(")


def test_pool_seed_matches_engine(tmp_path):
    result = _live_result()
    out = tmp_path / "live.xlsx"
    ElevateWorkbookBuilder(rates_path=RATES, gainsharing_result=result).build(out)
    pool = load_workbook(out)["Pool & Cash Gate"]
    p = result.pool_df.iloc[0]
    # Baseline seed on row 5, col 2.
    assert pool.cell(5, 2).value == pytest.approx(p["baseline_egp"])


def test_pipeline_all_stages_ok(tmp_path):
    import run_pipeline
    report = run_pipeline.run_pipeline(
        rates_path=MODULE_DIR / "target_rates.json",
        quote_path=MODULE_DIR / "sample_inputs" / "supplier_quote.txt",
        notes_path=MODULE_DIR / "sample_inputs" / "site_notes.json",
        out_dir=tmp_path,
    )
    assert all(v.startswith("OK") for v in report.values()), report
    assert (tmp_path / "UNITED_BROTHERS_ELEVATE_MASTER.xlsx").exists()
