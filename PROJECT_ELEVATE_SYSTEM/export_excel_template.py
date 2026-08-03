#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_excel_template.py
========================
United Brothers Co. / الاخوة المتحدين للمقاولات
PROJECT ELEVATE (Bulletproof Enterprise Edition)

Branded Excel Generator
مولّد قالب إكسل المؤسسي

Builds `UNITED_BROTHERS_ELEVATE_MASTER.xlsx` with openpyxl:
  * Corporate Navy (#1B365D) headers with white bold text.
  * Gold/Amber (#D4AF37) metric & total accents.
  * Live Excel formulas: SUM, IF, AND, COUNTA (self-recalculating workbook).
  * Dynamic conditional formatting (Approved / SLA-penalty / Disqualified palette).
  * Bilingual (Arabic RTL + English) headers, frozen panes, auto-fit columns,
    and an audit-ready governance reference sheet.

Author: AI Operations Architect — United Brothers Co.
Python: 3.10+
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side, NamedStyle,
    )
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "[FATAL] openpyxl is required. Install with: pip install openpyxl\n"
        "        مطلوب مكتبة openpyxl."
    ) from exc

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | ELEVATE-XLSX | %(levelname)-7s | %(message)s",
)
logger = logging.getLogger("elevate.xlsx")

DEFAULT_RATES_PATH = Path(__file__).with_name("target_rates.json")
OUTPUT_FILE = "UNITED_BROTHERS_ELEVATE_MASTER.xlsx"

# --------------------------------------------------------------------------- #
#  Corporate theme  |  الهوية المؤسسية
# --------------------------------------------------------------------------- #
NAVY = "1B365D"
GOLD = "D4AF37"
WHITE = "FFFFFF"

# Status palette (fill, font).
GREEN_FILL, GREEN_TEXT = "E8F5E9", "2E7D32"     # Full payout / Approved
AMBER_FILL, AMBER_TEXT = "FFF3E0", "EF6C00"     # SLA penalty / Warning
CRIMSON_FILL, CRIMSON_TEXT = "FFEBEE", "C62828"  # Disqualified / Hold / Crisis

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ENTITY_EN = "United Brothers Co."
ENTITY_AR = "الاخوة المتحدين للمقاولات"


# --------------------------------------------------------------------------- #
class ElevateWorkbookBuilder:
    def __init__(
        self,
        rates_path: str | Path = DEFAULT_RATES_PATH,
        gainsharing_result: Any = None,
    ) -> None:
        """
        rates_path: path to target_rates.json.
        gainsharing_result: optional GainsharingResult from
            gainsharing_calculator.GainsharingCalculator.run(). When provided,
            the Gainsharing and Pool sheets are seeded with the real computed
            figures; otherwise illustrative demo rows are used. Live Excel
            formulas are preserved in both cases so the workbook stays
            self-recalculating and audit-ready.
        """
        self.rates_path = Path(rates_path)
        self.config = self._load_config(self.rates_path)
        self.gov = self.config["governance"]
        self.result = gainsharing_result
        self.wb = Workbook()
        self._register_styles()

    # ------------------------------------------------------------------ #
    #  Data-source helpers  |  مصادر البيانات
    # ------------------------------------------------------------------ #
    def _gainsharing_rows(self) -> list[dict[str, Any]]:
        """Uniform per-member rows for the Gainsharing sheet (live or demo)."""
        if self.result is not None:
            rows: list[dict[str, Any]] = []
            for _, r in self.result.distribution_df.iterrows():
                rows.append({
                    "name": r.get("name", ""),
                    "role": r.get("role", ""),
                    "badge": r.get("ld_badge", "Level 1"),
                    "mult": float(r.get("badge_multiplier", 1.0)),
                    "tw": float(r.get("time_weight", 1.0)),
                    "ppc": float(r.get("ppc", 0.0)),
                    "sla": 1 if r.get("section_sla_met", True) else 0,
                    "ld": 1 if r.get("ld_sla_met", True) else 0,
                    "base": float(r.get("base_share_egp", 0.0)),
                    "perf": float(r.get("perf_share_egp", 0.0)),
                })
            return rows
        # Illustrative demo rows (base/perf left at 0 for manual entry).
        demo = [
            ("Ahmed Fathy", "Site Manager", "Level 3", 1.35, 1.0, 0.92, 1, 1),
            ("Mona Adel", "QA/QC Engineer", "Level 2", 1.20, 0.8, 0.88, 1, 0),
            ("Khaled Samir", "Foreman", "Level 1", 1.00, 1.0, 0.70, 1, 1),
            ("Sara Nabil", "Planner", "Level 2", 1.20, 0.5, 0.90, 1, 1),
        ]
        return [
            {"name": n, "role": ro, "badge": b, "mult": mu, "tw": tw,
             "ppc": p, "sla": s, "ld": l, "base": 0.0, "perf": 0.0}
            for (n, ro, b, mu, tw, p, s, l) in demo
        ]

    def _pool_seed(self) -> dict[str, float]:
        """Input seed values for the Pool & Cash Gate sheet (live or demo)."""
        if self.result is not None:
            p = self.result.pool_df.iloc[0]
            return {
                "baseline": float(p["baseline_egp"]),
                "delta": float(p["escalation_delta"]),
                "actual": float(p["actual_egp"]),
                "qf": float(p["quality_factor"]),
                "bad_debt": float(p["bad_debt_egp"]),
                "cash": float(p["cash_collected_pct"]),
            }
        return {"baseline": 12_000_000, "delta": 0.08, "actual": 10_400_000,
                "qf": 0.95, "bad_debt": 150_000, "cash": 0.82}

    @staticmethod
    def _load_config(path: Path) -> dict[str, Any]:
        if not path.exists():
            raise FileNotFoundError(f"[CONFIG] target_rates.json not found: {path}")
        with path.open("r", encoding="utf-8") as fh:
            return json.load(fh)

    # ------------------------------------------------------------------ #
    #  Styles  |  الأنماط
    # ------------------------------------------------------------------ #
    def _register_styles(self) -> None:
        header = NamedStyle(name="ub_header")
        header.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        header.fill = PatternFill("solid", fgColor=NAVY)
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header.border = BORDER

        title = NamedStyle(name="ub_title")
        title.font = Font(bold=True, color=WHITE, size=16, name="Calibri")
        title.fill = PatternFill("solid", fgColor=NAVY)
        title.alignment = Alignment(horizontal="center", vertical="center")

        gold = NamedStyle(name="ub_gold")
        gold.font = Font(bold=True, color=NAVY, size=11, name="Calibri")
        gold.fill = PatternFill("solid", fgColor=GOLD)
        gold.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        gold.border = BORDER

        money = NamedStyle(name="ub_money")
        money.number_format = '#,##0.00 "EGP"'
        money.alignment = Alignment(horizontal="right", vertical="center")
        money.border = BORDER

        pct = NamedStyle(name="ub_pct")
        pct.number_format = "0.0%"
        pct.alignment = Alignment(horizontal="center", vertical="center")
        pct.border = BORDER

        for style in (header, title, gold, money, pct):
            if style.name not in self.wb.named_styles:
                self.wb.add_named_style(style)

    # ------------------------------------------------------------------ #
    #  Shared header block  |  ترويسة مشتركة
    # ------------------------------------------------------------------ #
    def _title_block(self, ws: Worksheet, subtitle_en: str, subtitle_ar: str,
                     span: int) -> int:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
        c = ws.cell(row=1, column=1, value=f"{ENTITY_EN}  |  {ENTITY_AR}")
        c.style = "ub_title"
        ws.row_dimensions[1].height = 34

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        c2 = ws.cell(row=2, column=2 if False else 1,
                     value=f"PROJECT ELEVATE — {subtitle_en}  |  {subtitle_ar}")
        c2.font = Font(bold=True, color=NAVY, size=12)
        c2.fill = PatternFill("solid", fgColor=GOLD)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22
        return 4  # first data row

    def _write_headers(self, ws: Worksheet, row: int, headers: list[str]) -> None:
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.style = "ub_header"
        ws.row_dimensions[row].height = 30
        ws.freeze_panes = ws.cell(row=row + 1, column=1)

    @staticmethod
    def _autofit(ws: Worksheet, widths: dict[int, int]) -> None:
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

    # ------------------------------------------------------------------ #
    #  Sheet 1: Gainsharing Dashboard
    # ------------------------------------------------------------------ #
    def build_gainsharing_sheet(self) -> None:
        ws = self.wb.active
        ws.title = "Gainsharing"
        ws.sheet_view.rightToLeft = False  # LTR sheet; AR text still renders

        headers = [
            "Staff / الموظف",
            "Role / الوظيفة",
            "L&D Badge / الشارة",
            "Badge x / المضاعف",
            "Time Weight / الوزن الزمني",
            "PPC %",
            "Section SLA / الالتزام",
            "L&D/SLA Met / استيفاء",
            "Base Share / الحصة الأساسية",
            "Perf Share / حصة الأداء",
            "Gross Share / الإجمالي",
            "Immediate 70% / الفوري",
            "Retained 30% / المحتجز",
            "Status / الحالة",
        ]
        start = self._title_block(ws, "Gainsharing Distribution",
                                  "توزيع المشاركة في المكاسب", span=len(headers))
        self._write_headers(ws, start, headers)

        # Rows come from a live GainsharingResult when provided, else demo rows.
        # Live figures seed Base/Perf; the rest stays formula-driven so the
        # workbook recalculates in Excel and remains audit-ready.
        rows = self._gainsharing_rows()
        ppc_gate = float(self.gov["ppc_eligibility_threshold"])
        immediate = float(self.gov["immediate_payout_pct"])
        retained = float(self.gov["retained_cushion_pct"])

        first = start + 1
        for i, m in enumerate(rows):
            r = first + i
            ws.cell(row=r, column=1, value=m["name"]).border = BORDER
            ws.cell(row=r, column=2, value=m["role"]).border = BORDER
            ws.cell(row=r, column=3, value=m["badge"]).border = BORDER
            ws.cell(row=r, column=4, value=m["mult"]).border = BORDER
            ws.cell(row=r, column=5, value=m["tw"]).border = BORDER
            c_ppc = ws.cell(row=r, column=6, value=m["ppc"]); c_ppc.style = "ub_pct"
            ws.cell(row=r, column=7, value=m["sla"]).border = BORDER
            ws.cell(row=r, column=8, value=m["ld"]).border = BORDER

            # Base / Perf shares — seeded from the engine (0 in demo mode).
            base = ws.cell(row=r, column=9, value=m["base"]); base.style = "ub_money"
            perf = ws.cell(row=r, column=10, value=m["perf"]); perf.style = "ub_money"

            # Gross = Base + Perf  (live SUM formula)
            gross = ws.cell(row=r, column=11,
                            value=f"=SUM(I{r}:J{r})")
            gross.style = "ub_money"
            # Immediate 70% / Retained 30%
            imm = ws.cell(row=r, column=12, value=f"=K{r}*{immediate}")
            imm.style = "ub_money"
            ret = ws.cell(row=r, column=13, value=f"=K{r}*{retained}")
            ret.style = "ub_money"

            # Status: IF(AND(SLA, PPC>=gate, L&D)) => Approved / Penalty / Ineligible
            status = ws.cell(
                row=r, column=14,
                value=(
                    f'=IF(G{r}<>1,"DISQUALIFIED",'
                    f'IF(AND(F{r}>={ppc_gate},H{r}=1),"APPROVED",'
                    f'IF(F{r}<{ppc_gate},"INELIGIBLE_PPC","PENALTY")))'
                ),
            )
            status.border = BORDER
            status.alignment = Alignment(horizontal="center", vertical="center")

        last = first + len(rows) - 1

        # Totals row (Gold accent + SUM / COUNTA).
        tr = last + 1
        tcell = ws.cell(row=tr, column=1, value="TOTAL / الإجمالي")
        tcell.style = "ub_gold"
        ws.cell(row=tr, column=2, value=f'=COUNTA(A{first}:A{last})&" staff"').style = "ub_gold"
        for col in (9, 10, 11, 12, 13):
            cell = ws.cell(row=tr, column=col,
                           value=f"=SUM({get_column_letter(col)}{first}:{get_column_letter(col)}{last})")
            cell.number_format = '#,##0.00 "EGP"'
            cell.font = Font(bold=True, color=NAVY)
            cell.fill = PatternFill("solid", fgColor=GOLD)
            cell.border = BORDER

        # Conditional formatting on Status column (N).
        status_range = f"N{first}:N{last}"
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=['"APPROVED"'],
                       fill=PatternFill("solid", fgColor=GREEN_FILL),
                       font=Font(color=GREEN_TEXT, bold=True)),
        )
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=['"PENALTY"'],
                       fill=PatternFill("solid", fgColor=AMBER_FILL),
                       font=Font(color=AMBER_TEXT, bold=True)),
        )
        for bad in ('"DISQUALIFIED"', '"INELIGIBLE_PPC"'):
            ws.conditional_formatting.add(
                status_range,
                CellIsRule(operator="equal", formula=[bad],
                           fill=PatternFill("solid", fgColor=CRIMSON_FILL),
                           font=Font(color=CRIMSON_TEXT, bold=True)),
            )
        # Highlight PPC below gate in amber.
        ws.conditional_formatting.add(
            f"F{first}:F{last}",
            CellIsRule(operator="lessThan", formula=[str(ppc_gate)],
                       fill=PatternFill("solid", fgColor=AMBER_FILL),
                       font=Font(color=AMBER_TEXT)),
        )

        self._autofit(ws, {1: 18, 2: 16, 3: 14, 4: 12, 5: 14, 6: 8, 7: 14,
                           8: 12, 9: 18, 10: 16, 11: 18, 12: 16, 13: 16, 14: 18})

    # ------------------------------------------------------------------ #
    #  Sheet 2: Pool & Cash Gate
    # ------------------------------------------------------------------ #
    def build_pool_sheet(self) -> None:
        ws = self.wb.create_sheet("Pool & Cash Gate")
        headers = ["Parameter / المعامل", "Value / القيمة", "Notes / ملاحظات"]
        start = self._title_block(ws, "Pool & Cash Gate", "المجمع وبوابة النقد",
                                  span=len(headers))
        self._write_headers(ws, start, headers)

        r = start + 1

        def put(label: str, value: Any, note: str, money: bool = False,
                pct: bool = False) -> int:
            nonlocal r
            ws.cell(row=r, column=1, value=label).border = BORDER
            c = ws.cell(row=r, column=2, value=value)
            c.border = BORDER
            if money:
                c.number_format = '#,##0.00 "EGP"'
            if pct:
                c.number_format = "0.0%"
            ws.cell(row=r, column=3, value=note).border = BORDER
            cur = r
            r += 1
            return cur

        seed = self._pool_seed()
        row_base = put("Baseline Cost / التكلفة الأساسية", seed["baseline"], "C_baseline", money=True)
        row_delta = put("Escalation Δ / مؤشر التصاعد", seed["delta"], "Material index", pct=True)
        row_adj = put("Adjusted Baseline / الأساس المعدل",
                      f"=B{row_base}*(1+B{row_delta})", "C_baseline*(1+Δ)", money=True)
        row_actual = put("Actual Cost / التكلفة الفعلية", seed["actual"], "C_actual", money=True)
        row_qf = put("Quality Factor / معامل الجودة", seed["qf"], "F_quality 0-1", pct=True)
        row_bd = put("Bad Debt / الديون المعدومة", seed["bad_debt"], "Isolated", money=True)
        row_s = put("Net Savings S / صافي الوفورات",
                    f"=MAX(0,(B{row_adj}-B{row_actual}-B{row_bd}))*B{row_qf}",
                    "S = MAX(0, adj-actual-baddebt)*F_q", money=True)
        team_pct = float(self.gov["team_shared_pool_pct"])
        row_pool = put("Team Pool 35% / مجمع الفريق",
                       f"=B{row_s}*{team_pct}", "P_pool = S*0.35", money=True)
        row_cash = put("Cash Collected % / نسبة التحصيل", seed["cash"], "Cash gate input", pct=True)

        full = float(self.config["cash_gate"]["full_unlock_threshold"])
        floor = float(self.config["cash_gate"]["partial_unlock_floor"])
        denom = float(self.config["cash_gate"]["prorata_denominator"])
        row_unlock = put(
            "Unlock Ratio / نسبة الفتح",
            f'=IF(B{row_cash}>={full},1,IF(B{row_cash}>={floor},B{row_cash}/{denom},0))',
            "Tiered cash gate", pct=True,
        )
        row_unlocked = put("Unlocked Pool / المجمع المفتوح",
                           f"=B{row_pool}*B{row_unlock}", "raw*unlock", money=True)

        base_pct = float(self.gov["base_equal_pool_pct"])
        perf_pct = float(self.gov["performance_pool_pct"])
        put("Base Pool 80% / المجمع الأساسي", f"=B{row_unlocked}*{base_pct}",
            "80% equal pool", money=True)
        put("Perf Pool 20% / مجمع الأداء", f"=B{row_unlocked}*{perf_pct}",
            "20% performance", money=True)
        imm = float(self.gov["immediate_payout_pct"])
        ret = float(self.gov["retained_cushion_pct"])
        put("Immediate 70% / الفوري", f"=B{row_unlocked}*{imm}", "Paid on gate", money=True)
        put("Retained 30% / المحتجز", f"=B{row_unlocked}*{ret}", "Cushion", money=True)

        # Conditional formatting: cash gate status color on unlock ratio.
        ws.conditional_formatting.add(
            f"B{row_unlock}",
            CellIsRule(operator="greaterThanOrEqual", formula=["1"],
                       fill=PatternFill("solid", fgColor=GREEN_FILL),
                       font=Font(color=GREEN_TEXT, bold=True)),
        )
        ws.conditional_formatting.add(
            f"B{row_unlock}",
            CellIsRule(operator="between", formula=["0.0001", "0.9999"],
                       fill=PatternFill("solid", fgColor=AMBER_FILL),
                       font=Font(color=AMBER_TEXT, bold=True)),
        )
        ws.conditional_formatting.add(
            f"B{row_unlock}",
            CellIsRule(operator="equal", formula=["0"],
                       fill=PatternFill("solid", fgColor=CRIMSON_FILL),
                       font=Font(color=CRIMSON_TEXT, bold=True)),
        )
        self._autofit(ws, {1: 34, 2: 22, 3: 34})

    # ------------------------------------------------------------------ #
    #  Sheet 3: BOQ Audit
    # ------------------------------------------------------------------ #
    def build_boq_sheet(self) -> None:
        ws = self.wb.create_sheet("BOQ Audit")
        headers = [
            "Item Code / الكود", "Description / الوصف", "Unit / الوحدة",
            "Qty / الكمية", "Quoted EGP / السعر", "Target EGP / المستهدف",
            "PPV/Unit / الانحراف", "PPV Total / إجمالي الانحراف",
            "Line Total / إجمالي البند", "Approved VO / اعتماد",
            "Scope Flag / حالة النطاق",
        ]
        start = self._title_block(ws, "BOQ & Quote Audit", "تدقيق جدول الكميات",
                                  span=len(headers))
        self._write_headers(ws, start, headers)

        threshold = float(self.gov["unapproved_scope_threshold_egp"])
        rows = []
        for tr in self.config.get("target_rates", []):
            rows.append((tr["item_code"], tr["description_en"], tr["unit"],
                         tr["target_rate_egp"]))

        first = start + 1
        for i, (code, desc, unit, target) in enumerate(rows):
            r = first + i
            ws.cell(row=r, column=1, value=code).border = BORDER
            ws.cell(row=r, column=2, value=desc).border = BORDER
            ws.cell(row=r, column=3, value=unit).border = BORDER
            ws.cell(row=r, column=4, value=1).border = BORDER   # qty (editable)
            q = ws.cell(row=r, column=5, value=target); q.number_format = '#,##0.00'
            t = ws.cell(row=r, column=6, value=target); t.number_format = '#,##0.00'
            # PPV per unit = Quoted - Target
            ppvu = ws.cell(row=r, column=7, value=f"=E{r}-F{r}")
            ppvu.number_format = '#,##0.00'
            # PPV total = PPV/unit * qty
            ppvt = ws.cell(row=r, column=8, value=f"=G{r}*D{r}")
            ppvt.number_format = '#,##0.00'
            # Line total
            lt = ws.cell(row=r, column=9, value=f"=E{r}*D{r}")
            lt.number_format = '#,##0.00 "EGP"'
            ws.cell(row=r, column=10, value=0).border = BORDER  # approved VO (0/1)
            # Scope flag: IF(AND(LineTotal>threshold, VO<>1),"UNAPPROVED","OK")
            flag = ws.cell(
                row=r, column=11,
                value=f'=IF(AND(I{r}>{threshold},J{r}<>1),"UNAPPROVED","OK")',
            )
            flag.alignment = Alignment(horizontal="center")
            flag.border = BORDER

        last = first + len(rows) - 1
        # Totals
        tr_row = last + 1
        ws.cell(row=tr_row, column=1, value="TOTAL / الإجمالي").style = "ub_gold"
        for col in (8, 9):
            cell = ws.cell(row=tr_row, column=col,
                           value=f"=SUM({get_column_letter(col)}{first}:{get_column_letter(col)}{last})")
            cell.number_format = '#,##0.00 "EGP"'
            cell.fill = PatternFill("solid", fgColor=GOLD)
            cell.font = Font(bold=True, color=NAVY)
            cell.border = BORDER

        # Conditional format: PPV Total overspend (>0) crimson, savings (<0) green.
        ws.conditional_formatting.add(
            f"H{first}:H{last}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=PatternFill("solid", fgColor=CRIMSON_FILL),
                       font=Font(color=CRIMSON_TEXT)),
        )
        ws.conditional_formatting.add(
            f"H{first}:H{last}",
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=PatternFill("solid", fgColor=GREEN_FILL),
                       font=Font(color=GREEN_TEXT)),
        )
        # Scope flag coloring.
        ws.conditional_formatting.add(
            f"K{first}:K{last}",
            FormulaRule(formula=[f'K{first}="UNAPPROVED"'],
                        fill=PatternFill("solid", fgColor=CRIMSON_FILL),
                        font=Font(color=CRIMSON_TEXT, bold=True)),
        )
        self._autofit(ws, {1: 18, 2: 34, 3: 10, 4: 10, 5: 14, 6: 14, 7: 14,
                           8: 16, 9: 18, 10: 12, 11: 16})

    # ------------------------------------------------------------------ #
    #  Sheet 4: Site KPIs
    # ------------------------------------------------------------------ #
    def build_site_sheet(self) -> None:
        ws = self.wb.create_sheet("Site KPIs")
        headers = [
            "Site / الموقع", "Date / التاريخ", "PPC %", "NCR", "OEE %",
            "LTI", "PPC Gate / بوابة", "Safety Gate / السلامة",
        ]
        start = self._title_block(ws, "Daily Site KPIs", "مؤشرات الموقع اليومية",
                                  span=len(headers))
        self._write_headers(ws, start, headers)

        ppc_gate = float(self.gov["ppc_eligibility_threshold"])
        oee_gate = float(self.gov["oee_uptime_threshold"])

        sample = [
            ("New Capital Tower B", "2026-08-02", 0.88, 1, 0.96, 0),
            ("Suez Pipe Rack", "2026-08-02", 0.82, 0, 0.94, 0),
            ("Ain Sokhna Warehouse", "2026-08-02", 0.79, 2, 0.93, 1),
        ]
        first = start + 1
        for i, (site, date, ppc, ncr, oee, lti) in enumerate(sample):
            r = first + i
            ws.cell(row=r, column=1, value=site).border = BORDER
            ws.cell(row=r, column=2, value=date).border = BORDER
            c1 = ws.cell(row=r, column=3, value=ppc); c1.style = "ub_pct"
            ws.cell(row=r, column=4, value=ncr).border = BORDER
            c2 = ws.cell(row=r, column=5, value=oee); c2.style = "ub_pct"
            ws.cell(row=r, column=6, value=lti).border = BORDER
            # PPC gate check
            g1 = ws.cell(row=r, column=7,
                         value=f'=IF(C{r}>={ppc_gate},"OK","BELOW")')
            g1.alignment = Alignment(horizontal="center"); g1.border = BORDER
            # Safety gate check
            g2 = ws.cell(row=r, column=8,
                         value=f'=IF(F{r}=0,"OK","DISQUALIFIED")')
            g2.alignment = Alignment(horizontal="center"); g2.border = BORDER

        last = first + len(sample) - 1
        # Conditional formats.
        ws.conditional_formatting.add(
            f"C{first}:C{last}",
            CellIsRule(operator="lessThan", formula=[str(ppc_gate)],
                       fill=PatternFill("solid", fgColor=AMBER_FILL),
                       font=Font(color=AMBER_TEXT)),
        )
        ws.conditional_formatting.add(
            f"E{first}:E{last}",
            CellIsRule(operator="lessThan", formula=[str(oee_gate)],
                       fill=PatternFill("solid", fgColor=AMBER_FILL),
                       font=Font(color=AMBER_TEXT)),
        )
        ws.conditional_formatting.add(
            f"F{first}:F{last}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=PatternFill("solid", fgColor=CRIMSON_FILL),
                       font=Font(color=CRIMSON_TEXT, bold=True)),
        )
        ws.conditional_formatting.add(
            f"H{first}:H{last}",
            FormulaRule(formula=[f'H{first}="DISQUALIFIED"'],
                        fill=PatternFill("solid", fgColor=CRIMSON_FILL),
                        font=Font(color=CRIMSON_TEXT, bold=True)),
        )
        self._autofit(ws, {1: 24, 2: 14, 3: 8, 4: 6, 5: 8, 6: 6, 7: 16, 8: 18})

    # ------------------------------------------------------------------ #
    #  Sheet 5: Governance reference
    # ------------------------------------------------------------------ #
    def build_governance_sheet(self) -> None:
        ws = self.wb.create_sheet("Governance")
        headers = ["Rule / القاعدة", "Value / القيمة"]
        start = self._title_block(ws, "Governance Reference", "مرجع الحوكمة",
                                  span=len(headers))
        self._write_headers(ws, start, headers)

        items = [
            ("Team Shared Pool / مجمع الفريق", f"{self.gov['team_shared_pool_pct']:.0%}"),
            ("Company Retention / احتفاظ الشركة", f"{self.gov['company_retention_pct']:.0%}"),
            ("Base Equal Pool / المجمع الأساسي", f"{self.gov['base_equal_pool_pct']:.0%}"),
            ("Performance Pool / مجمع الأداء", f"{self.gov['performance_pool_pct']:.0%}"),
            ("Immediate Payout / الدفع الفوري", f"{self.gov['immediate_payout_pct']:.0%}"),
            ("Retained Cushion / الوسادة المحتجزة", f"{self.gov['retained_cushion_pct']:.0%}"),
            ("Subk Back-charge Reserve / احتياطي المقاول", f"{self.gov['subcontractor_backcharge_reserve_pct']:.0%}"),
            ("Unapproved Scope Threshold / حد النطاق", f"{self.gov['unapproved_scope_threshold_egp']:,.0f} EGP"),
            ("Scope Penalty / غرامة النطاق", f"{self.gov['unapproved_scope_penalty_pct']:.0%}"),
            ("PPC Eligibility / أهلية الإنجاز", f"{self.gov['ppc_eligibility_threshold']:.0%}"),
            ("OEE Uptime Gate / بوابة الكفاءة", f"{self.gov['oee_uptime_threshold']:.0%}"),
            ("Fast-track VO / تسوية سريعة", f"< {self.gov['fast_track_vo_days']} days"),
            ("VE Reward / مكافأة الهندسة القيمية", f"{self.gov['value_engineering_reward_pct']:.0%}"),
            ("L&D Forfeit / مصادرة التعلم", f"{self.gov['ld_forfeit_pct']:.0%}"),
            ("Zero LTI Gate / السلامة", "0 LTI — mandatory"),
        ]
        first = start + 1
        for i, (rule, val) in enumerate(items):
            r = first + i
            ws.cell(row=r, column=1, value=rule).border = BORDER
            c = ws.cell(row=r, column=2, value=val)
            c.font = Font(bold=True, color=NAVY)
            c.alignment = Alignment(horizontal="center")
            c.border = BORDER
        self._autofit(ws, {1: 42, 2: 20})

    # ------------------------------------------------------------------ #
    def build(self, output_path: str | Path = OUTPUT_FILE) -> Path:
        logger.info("Building workbook...")
        self.build_gainsharing_sheet()
        self.build_pool_sheet()
        self.build_boq_sheet()
        self.build_site_sheet()
        self.build_governance_sheet()

        # Workbook-level metadata (audit readiness).
        self.wb.properties.creator = ENTITY_EN
        self.wb.properties.title = "PROJECT ELEVATE — Master Workbook"
        self.wb.properties.company = ENTITY_EN

        out = Path(output_path)
        self.wb.save(out)
        logger.info("Saved workbook: %s", out.resolve())
        return out


def _demo_gainsharing_result(rates_path: str | Path):
    """Run the gainsharing engine on the demo scenario and return its result.

    Imported lazily so this module still builds a blank template even if
    gainsharing_calculator's dependencies are unavailable.
    """
    from gainsharing_calculator import (
        GainsharingCalculator, ProjectFinancials, TeamMember,
    )

    calc = GainsharingCalculator(rates_path=rates_path)
    fin = ProjectFinancials(
        project_name="Ain Sokhna Industrial Warehouse",
        baseline_cost_egp=12_000_000.0, actual_cost_egp=10_400_000.0,
        cash_collected_pct=0.82, quality_factor=0.95,
        escalation_commodity="steel_rebar", bad_debt_egp=150_000.0,
        subcontractor_value_egp=3_000_000.0, lost_time_injuries=0,
    )
    members = [
        TeamMember("Ahmed Fathy", "Site Manager", time_weight=1.0, ld_badge="Level 3",
                   ppc=0.92, equipment_oee=0.97, vo_settlement_days=5,
                   value_engineering_savings_egp=200_000),
        TeamMember("Mona Adel", "QA/QC Engineer", time_weight=0.8, ld_badge="Level 2",
                   ppc=0.88, ld_sla_met=False),
        TeamMember("Khaled Samir", "Foreman", time_weight=1.0, ld_badge="Level 1", ppc=0.70),
        TeamMember("Sara Nabil", "Planner", time_weight=0.5, ld_badge="Level 2",
                   ppc=0.90, resigning_clean_handover=True, kaizen_points=2),
    ]
    return calc.run(fin, members)


def _main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="United Brothers Co. Excel Template Generator")
    parser.add_argument("--rates", type=str, default=str(DEFAULT_RATES_PATH))
    parser.add_argument("--out", type=str, default=OUTPUT_FILE)
    parser.add_argument("--live", action="store_true",
                        help="Seed the Gainsharing & Pool sheets from a live "
                             "gainsharing_calculator run instead of demo rows.")
    args = parser.parse_args()

    result = _demo_gainsharing_result(args.rates) if args.live else None
    builder = ElevateWorkbookBuilder(rates_path=args.rates, gainsharing_result=result)
    path = builder.build(args.out)
    mode = "LIVE (engine-seeded)" if args.live else "TEMPLATE (demo rows)"
    print(f"✅ Workbook generated [{mode}]: {path.resolve()}")


if __name__ == "__main__":
    _main()
